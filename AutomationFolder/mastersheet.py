import os
import pandas as pd
from openpyxl import load_workbook

def genmasterSheet(directory, savePath):
    results = []

    # Scan all files in the directory
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            # Construct the full file path
            file_path = os.path.join(directory, filename)
            
            # Load the workbook
            wb = load_workbook(file_path)
            
            # Select the active worksheet
            ws = wb.active
            
            # Initialize the sum with the second value of the third column
            sum_values = ws.cell(row=2, column=3).value
            
            # Iterate over the rest of the values in the third column
            for i in range(3, ws.max_row + 1):
                sum_values += ws.cell(row=i, column=3).value
            
            # Extract the last value in column A that contains a value
            df = pd.read_excel(file_path)
            last_value_in_A = df['Label'].dropna().iloc[-1]
            
            # Create a new DataFrame with the desired column order
            results.append(pd.DataFrame({'File Location': [file_path], 
                                         'Number of Objects': [last_value_in_A], 
                                         'Sum of Integrated Density': [sum_values]}))

    # Concatenate all DataFrames in the list
    results = pd.concat(results, ignore_index=True)

    # Write the results to a new Excel file
    results.to_excel(savePath, index=False)
    print("Done summary file generation")


